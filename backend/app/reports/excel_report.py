from __future__ import annotations

from datetime import datetime
from pathlib import Path
import json

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPORT_DIR = Path("/app/reports")
REPORT_DIR.mkdir(parents=True, exist_ok=True)

BG = "F6F1E6"
CARD = "FFFDF7"
CARD_ALT = "FFF4D6"
INK = "17120D"
MUTED = "61584D"
RED_DARK = "65000B"
RED = "B00020"
GOLD = "B8871B"
GOLD_LIGHT = "FFE2A0"
GREEN = "1F8F5F"
ORANGE = "D06B00"
BORDER = "D7BB74"
WHITE = "FFFFFF"

SEVERITY_FILL = {
    "critical": "65000B",
    "high": "B00020",
    "medium": "B8871B",
    "low": "7B5500",
    "info": "61584D",
}


def generate_excel_report(audit: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthese Direction"

    _sheet_summary(ws, audit)
    _sheet_dnsdumpster_like(wb, audit)
    _sheet_action_plan(wb, audit)
    _sheet_findings(wb, audit)
    _sheet_exposure(wb, audit)
    _sheet_nmap(wb, audit)
    _sheet_tls_mail_dns(wb, audit)
    _sheet_cti_sla(wb, audit)
    _sheet_graph(wb, audit)
    _sheet_limits(wb, audit)
    _sheet_raw(wb, audit)

    filename = f"open_easm_beta_{audit['domain'].replace('.', '_')}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = REPORT_DIR / filename
    wb.save(path)
    return filename


def _sheet_summary(ws, audit: dict):
    _paint(ws, 42, 12)
    ws.merge_cells("A1:L3")
    ws["A1"] = "OPENEASM BETA — RAPPORT D'EXPOSITION EXTERNE"
    ws["A1"].font = Font(bold=True, size=19, color=GOLD_LIGHT)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=RED_DARK)

    score = audit.get("score", {}) or {}
    risk = audit.get("executive_risk", {}) or {}
    scan = audit.get("service_scan", {}) or {}
    ip = audit.get("ip_inventory", {}) or {}
    sub = audit.get("subdomains", {}) or {}
    profile = audit.get("domain_profile", {}) or {}

    kpis = [
        ("Domaine", audit.get("domain")),
        ("Score global", f"{score.get('score', 'N/A')} / {score.get('max_score', 1000)}"),
        ("Niveau", score.get("level")),
        ("Risque exécutif", f"{risk.get('overall_score', 'N/A')} / {risk.get('max_score', 100)}"),
        ("Posture", risk.get("posture")),
        ("Profil", profile.get("label")),
        ("IP publiques", ip.get("public_ip_count", 0)),
        ("Hosts canoniques", len((audit.get("canonical_hosts") or (audit.get("dnsdumpster_like", {}) or {}).get("canonical_hosts") or []))),
        ("Aliases regroupés", (audit.get("dnsdumpster_like", {}) or {}).get("aliases_count", 0)),
        ("Ports Nmap", len(scan.get("canonical_open_ports", scan.get("open_ports", [])) or [])),
        ("CVE service/version", scan.get("count_cves", 0)),
        ("TLS", f"{audit.get('tls_score', {}).get('global_score', 'N/A')} / 100"),
        ("Date audit", audit.get("created_at")),
    ]
    row, col = 5, 1
    for label, value in kpis:
        _metric_card(ws, row, col, label, value)
        col += 3
        if col > 10:
            col = 1
            row += 4

    ws.merge_cells("A17:L23")
    ws["A17"] = risk.get("board_summary") or _conclusion(audit)
    ws["A17"].font = Font(size=12, color=INK)
    ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A17"].fill = PatternFill("solid", fgColor=CARD)

    sev = score.get("by_severity", {}) or {}
    start = 26
    ws.cell(start, 1, "Répartition des constats")
    ws.cell(start, 1).font = Font(bold=True, color=RED_DARK, size=14)
    ws.cell(start + 1, 1, "Criticité")
    ws.cell(start + 1, 2, "Nombre")
    _header_row(ws, start + 1, 2)
    for idx, key in enumerate(["critical", "high", "medium", "low", "info"], start + 2):
        ws.cell(idx, 1, key)
        ws.cell(idx, 2, sev.get(key, 0))
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=SEVERITY_FILL.get(key, MUTED))
        ws.cell(idx, 1).font = Font(bold=True, color=WHITE)
        ws.cell(idx, 2).fill = PatternFill("solid", fgColor=CARD)

    chart = BarChart()
    chart.title = "Constats par criticité"
    chart.y_axis.title = "Nombre"
    chart.x_axis.title = "Criticité"
    data = Reference(ws, min_col=2, min_row=start + 1, max_row=start + 6)
    cats = Reference(ws, min_col=1, min_row=start + 2, max_row=start + 6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    ws.add_chart(chart, "D26")

    ws.merge_cells("A36:L40")
    ws["A36"] = "Portée Beta : rapport orienté décision et correction opérationnelle. La détection Nmap reste non exploitante. Les versions non exposées ne génèrent pas de CVE inventée."
    ws["A36"].font = Font(color=MUTED, italic=True)
    ws["A36"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A36"].fill = PatternFill("solid", fgColor=CARD_ALT)

    _format(ws, freeze=None)



def _sheet_dnsdumpster_like(wb, audit):
    dd = audit.get("dnsdumpster_like", {}) or {}
    summary = wb.create_sheet("DNSDumpster-like Summary")
    summary_rows = [
        ["Domain", dd.get("domain") or audit.get("domain")],
        ["System Locations", len(dd.get("system_locations", {}) or {})],
        ["Hosting / Networks", len(dd.get("hosting_networks", []) or [])],
        ["Services / Banners", len(dd.get("services_banners", {}) or {})],
        ["A Records enrichis", len(dd.get("a_records", []) or [])],
        ["MX Records enrichis", len(dd.get("mx_records", []) or [])],
        ["NS Records enrichis", len(dd.get("ns_records", []) or [])],
        ["TXT Records", len(dd.get("txt_records", []) or [])],
        ["Note", "Nmap est une source complémentaire ; HTTP/HTTPS/TLS sont intégrés même si Nmap est vide."],
    ]
    _table_sheet(summary, "DNSDUMPSTER-LIKE SUMMARY", ["Indicateur", "Valeur"], summary_rows)

    ws = wb.create_sheet("System Locations")
    rows = [[country, count] for country, count in (dd.get("system_locations", {}) or {}).items()]
    _table_sheet(ws, "SYSTEM LOCATIONS", ["Country", "Count"], rows or [["Unknown", 0]])

    ws = wb.create_sheet("Hosting Networks")
    rows = [[h.get("asn"), h.get("network"), h.get("asn_name"), h.get("country"), h.get("count")] for h in dd.get("hosting_networks", [])]
    _table_sheet(ws, "HOSTING / NETWORKS", ["ASN", "Network", "ASN Name", "Country", "Hosts"], rows)

    ws = wb.create_sheet("Services Banners")
    rows = [[banner, count] for banner, count in (dd.get("services_banners", {}) or {}).items()]
    _table_sheet(ws, "SERVICES / BANNERS", ["Banner", "Count"], rows)

    for title, key, headers, fn in [
        ("A Records", "a_records", _host_headers(), _host_row),
        ("MX Records", "mx_records", ["Priority"] + _host_headers(), lambda h: [h.get("priority")] + _host_row(h)),
        ("NS Records", "ns_records", _host_headers(), _host_row),
        ("Host Inventory", "hosts", _host_headers(), _host_row),
    ]:
        ws = wb.create_sheet(title)
        _table_sheet(ws, title.upper(), headers, [fn(h) for h in dd.get(key, [])])

    ws = wb.create_sheet("TXT Records")
    rows = []
    for item in dd.get("txt_records", []) or []:
        spf = item.get("spf") or {}
        rows.append([item.get("type"), item.get("value"), "\n".join(spf.get("includes", [])), "\n".join(spf.get("ip4", [])), "\n".join(spf.get("ip6", [])), "\n".join(spf.get("providers", [])), ", ".join(item.get("sources", []))])
    _table_sheet(ws, "TXT RECORDS", ["Type", "Value", "SPF Includes", "SPF ip4", "SPF ip6", "Providers", "Sources"], rows)


def _host_headers():
    return ["Host", "Aliases", "IP", "ASN", "Network", "ASN Name", "Country", "Provider", "Open Services", "HTTP Status", "HTTPS Status", "Final URL", "Title", "Server Banner", "TLS CN", "TLS SAN", "TLS Issuer", "TLS Expiration", "Technologies", "Nmap Services", "RevIP", "Sources"]


def _host_row(h):
    http = h.get("http") or {}
    https = h.get("https") or {}
    tls = h.get("tls") or {}
    services = _services_text(h.get("open_services", [])) or "Non détecté"
    server_banner = "; ".join(dict.fromkeys([s.get("banner") for s in h.get("open_services", []) if s.get("banner")]))
    title = https.get("title") or http.get("title") or next((s.get("title") for s in h.get("open_services", []) if s.get("title") and s.get("title") != "Non détecté"), "Non détecté")
    return [
        h.get("host"), "\n".join(h.get("aliases", []) or []), h.get("ip"), h.get("asn"), h.get("network"), h.get("asn_name"), h.get("country"), h.get("provider"),
        services, http.get("status_code") or "Non détecté", https.get("status_code") or "Non détecté", https.get("final_url") or http.get("final_url") or "Non détecté", title,
        server_banner or "Non détecté", tls.get("cn") or "Non détecté", "\n".join(tls.get("san", []) or []), tls.get("issuer") or "Non détecté", tls.get("expires_at") or "Non détecté",
        _tech_text(h.get("technologies", [])) or "Non détecté", _nmap_text(h.get("nmap_services", [])) or "Non détecté par Nmap", h.get("revip_count", 0), ", ".join(h.get("sources", [])),
    ]


def _services_text(services):
    parts = []
    for s in services or []:
        seg = f"{s.get('scheme')}: {s.get('banner') or s.get('service') or 'unknown server'}"
        if s.get("title") and s.get("title") != "Non détecté":
            seg += f"; title: {s.get('title')}"
        parts.append(seg)
    return "\n".join(parts)


def _tech_text(technologies):
    out = []
    for t in technologies or []:
        name = t.get("name") or ""
        if t.get("version"):
            name += f":{t.get('version')}"
        if t.get("note"):
            name += f" ({t.get('note')})"
        out.append(name)
    return "\n".join(dict.fromkeys(out))


def _nmap_text(services):
    out = []
    for s in services or []:
        version = s.get("version") or "Version non exposée"
        out.append(f"{s.get('port')}/{s.get('protocol', 'tcp')} {s.get('name') or s.get('service')} {s.get('product') or ''} {version}".strip())
    return "\n".join(out)

def _sheet_action_plan(wb, audit):
    ws = wb.create_sheet("Plan Action")
    rows = []
    for f in sorted(audit.get("findings", []), key=lambda x: _severity_order(x.get("severity", "info"))):
        sev = f.get("severity", "info")
        rows.append([
            _priority_label(sev), sev, f.get("category", ""), _loc(f.get("location", {})),
            f.get("title", ""), f.get("recommendation", ""), _sla_for(sev), "À qualifier",
        ])
    _table_sheet(ws, "PLAN D'ACTION PRIORISÉ", ["Priorité", "Sévérité", "Catégorie", "Lieu / Source", "Constat", "Action recommandée", "SLA cible", "Statut"], rows or [["Info", "info", "Aucun", "N/A", "Aucun constat", "Maintenir la surveillance", "Suivi", "N/A"]])


def _sheet_findings(wb, audit):
    ws = wb.create_sheet("Constats")
    rows = []
    for f in sorted(audit.get("findings", []), key=lambda x: _severity_order(x.get("severity", "info"))):
        loc = f.get("location", {}) or {}
        rows.append([
            f.get("severity"), f.get("category"), _loc(loc), loc.get("hostname"), loc.get("control"),
            f.get("title"), f.get("description"), f.get("recommendation"), ", ".join(f.get("applies_to", [])),
        ])
    _table_sheet(ws, "CONSTATS LOCALISÉS", ["Sévérité", "Catégorie", "Lieu / Source", "Hostname", "Contrôle", "Titre", "Description", "Recommandation", "Applicabilité"], rows)


def _sheet_exposure(wb, audit):
    ws = wb.create_sheet("Exposition")
    rows = []
    for item in audit.get("ip_inventory", {}).get("unique_ips", audit.get("ip_inventory", {}).get("display_ips", [])):
        rows.append([item.get("ip"), item.get("is_public"), item.get("scope"), ", ".join(item.get("sources", [])), "\n".join(item.get("hostnames", [])[:12])])
    _table_sheet(ws, "INVENTAIRE IP PUBLIC", ["IP", "Publique", "Périmètre", "Sources", "Hostnames"], rows)

    start = ws.max_row + 3
    ws.cell(start, 1, "Sous-domaines publics")
    ws.cell(start, 1).font = Font(bold=True, size=14, color=RED_DARK)
    sub_rows = [["Sous-domaine", "Source"]]
    sub = audit.get("subdomains", {}) or {}
    for name in sub.get("subdomains", [])[:250]:
        sub_rows.append([name, sub.get("source", "passif")])
    _write_table(ws, start + 2, ["Sous-domaine", "Source"], sub_rows[1:] or [["Aucun", sub.get("source", "passif")]])
    _format(ws)


def _sheet_nmap(wb, audit):
    ws = wb.create_sheet("Nmap Services")
    scan = audit.get("service_scan", {}) or {}
    rows = []
    for port in scan.get("canonical_open_ports", scan.get("open_ports", [])):
        cves = port.get("cves", []) or []
        if cves:
            for cve in cves:
                rows.append([
                    port.get("hostname"), port.get("port"), port.get("protocol"), port.get("name"), port.get("product"),
                    port.get("version") or "Version non exposée", "\n".join(port.get("cpe", [])), cve.get("cve"),
                    cve.get("severity"), cve.get("cvss"), cve.get("confidence"), cve.get("evidence"),
                ])
        else:
            rows.append([
                port.get("hostname"), port.get("port"), port.get("protocol"), port.get("name"), port.get("product"),
                port.get("version") or "Version non exposée", "\n".join(port.get("cpe", [])), "", "", "", "", port.get("evidence"),
            ])
    _table_sheet(ws, "NMAP SERVICE / VERSION / CVE — NON EXPLOITANT", ["Hostname", "Port", "Proto", "Service", "Produit", "Version", "CPE", "CVE", "Sévérité", "CVSS", "Confiance", "Preuve"], rows or [["Aucun", "", "", "", "", "", "", "", "", "", "", scan.get("note", "Aucun port détecté")]])

    start = ws.max_row + 3
    meta = [
        ["Mode", scan.get("mode")],
        ["Politique", scan.get("command_policy")],
        ["Durée", f"{scan.get('elapsed_seconds', 0)} s"],
        ["Note", scan.get("note")],
        ["Limite", "Aucune exploitation, aucun bruteforce, aucun DoS, aucun script NSE intrusif."],
    ]
    _write_table(ws, start, ["Métadonnée", "Valeur"], meta)
    _format(ws)


def _sheet_tls_mail_dns(wb, audit):
    ws = wb.create_sheet("DNS Mail TLS")
    rows = []
    for typ, data in (audit.get("dns", {}).get("records", {}) or {}).items():
        rows.append(["DNS", typ, data.get("status"), "\n".join(data.get("values", [])), data.get("error")])
    rows.append(["Mail", "MX", "", "\n".join(audit.get("mail", {}).get("mx", {}).get("values", [])), ""])
    rows.append(["Mail", "DMARC", "", "\n".join(audit.get("mail", {}).get("dmarc_records", [])), ""])
    for t in audit.get("tls_score", {}).get("targets", []):
        rows.append(["TLS", t.get("hostname"), t.get("level"), f"{t.get('score')} / 100 | {t.get('tls_version')} | expiration {t.get('days_remaining')} j", "\n".join(t.get("checks", []))])
    _table_sheet(ws, "DNS / MESSAGERIE / TLS", ["Famille", "Contrôle", "Statut", "Valeur", "Note"], rows)


def _sheet_cti_sla(wb, audit):
    ws = wb.create_sheet("CTI et SLA")
    rows = []
    for ipr in audit.get("cti", {}).get("ip_reputation_all", audit.get("cti", {}).get("ip_reputation", [])):
        for check in ipr.get("checks", []):
            rows.append([ipr.get("ip"), ipr.get("scope"), ", ".join(ipr.get("hostnames", [])), check.get("zone"), check.get("status"), ", ".join(check.get("values", [])) if check.get("values") else check.get("error") or check.get("detail") or ""])
    _table_sheet(ws, "CTI / RÉPUTATION", ["IP", "Périmètre", "Hostnames", "Zone", "Statut", "Détail"], rows)

    start = ws.max_row + 3
    sla_rows = []
    for item in audit.get("patching_sla", {}).get("items", []):
        sla_rows.append([item.get("id"), item.get("severity"), item.get("category"), item.get("title"), item.get("sla_days"), item.get("due_at"), item.get("status")])
    _write_table(ws, start, ["ID", "Sévérité", "Catégorie", "Constat", "SLA jours", "Échéance", "Statut"], sla_rows or [["N/A", "info", "N/A", "Aucun", "", "", "N/A"]])
    _format(ws)


def _sheet_graph(wb, audit):
    ws = wb.create_sheet("Graph Explorer")
    graph = audit.get("attack_graph", {}) or {}
    metrics = graph.get("metrics", {}) or {}
    rows = [["Nœuds", metrics.get("nodes", 0)], ["Relations", metrics.get("edges", 0)], ["Commentaire", "Cartographie relationnelle disponible dans l'onglet Graph Explorer de l'application."]]
    _table_sheet(ws, "GRAPH EXPLORER — SYNTHÈSE", ["Indicateur", "Valeur"], rows)


def _sheet_limits(wb, audit):
    ws = wb.create_sheet("Portee Limites")
    rows = [
        ["Nature", "Audit public défensif d'exposition externe."],
        ["Nmap", "Service/version/port uniquement. Aucun exploit, bruteforce, DoS ou script intrusif."],
        ["CVE", "Corrélation par version exposée. Une version masquée ne doit pas générer de faux positif."],
        ["Backports", "Une version apparente peut être corrigée par backport de sécurité côté distribution."],
        ["Responsabilité", "L'utilisateur doit disposer d'un droit, d'une autorisation explicite ou d'un motif légitime."],
    ]
    _table_sheet(ws, "PORTÉE, LIMITES ET RESPONSABILITÉ", ["Point", "Détail"], rows)


def _sheet_raw(wb, audit):
    ws = wb.create_sheet("Raw JSON Summary")
    rows = [
        ["id", audit.get("id")],
        ["domain", audit.get("domain")],
        ["created_at", audit.get("created_at")],
        ["mode", audit.get("mode")],
        ["score", json.dumps(audit.get("score", {}), ensure_ascii=False, default=str)],
        ["executive_risk", json.dumps(audit.get("executive_risk", {}), ensure_ascii=False, default=str)[:32000]],
        ["dnsdumpster_like", json.dumps(audit.get("dnsdumpster_like", {}), ensure_ascii=False, default=str)[:32000]],
        ["canonical_hosts", json.dumps(audit.get("canonical_hosts", []), ensure_ascii=False, default=str)[:32000]],
        ["raw_subdomains", json.dumps(audit.get("raw_subdomains", []), ensure_ascii=False, default=str)[:32000]],
        ["raw_nmap_services", json.dumps(audit.get("raw_nmap_services", []), ensure_ascii=False, default=str)[:32000]],
    ]
    _table_sheet(ws, "RÉSUMÉ JSON", ["Clé", "Valeur"], rows)


def _metric_card(ws, row, col, label, value):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    ws.cell(row, col, label)
    ws.cell(row + 1, col, _excel_value(value))
    ws.cell(row, col).font = Font(bold=True, color=GOLD, size=9)
    ws.cell(row + 1, col).font = Font(bold=True, color=INK, size=13)
    ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_ALT)
    ws.cell(row + 1, col).fill = PatternFill("solid", fgColor=CARD)
    ws.cell(row, col).alignment = Alignment(horizontal="center")
    ws.cell(row + 1, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _table_sheet(ws, title, headers, rows):
    _paint(ws, rows=max(len(rows) + 10, 24), cols=max(len(headers), 8))
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=max(len(headers), 8))
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color=GOLD_LIGHT)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=RED_DARK)
    _write_table(ws, 4, headers, rows)
    _format(ws)


def _write_table(ws, start_row, headers, rows):
    for col, h in enumerate(headers, 1):
        ws.cell(start_row, col, h)
    _header_row(ws, start_row, len(headers))
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx, _excel_value(value))
            cell.font = Font(color=INK)
            cell.fill = PatternFill("solid", fgColor=CARD if r_idx % 2 else ROW_FILL())
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if str(value).lower() in SEVERITY_FILL:
                cell.fill = PatternFill("solid", fgColor=SEVERITY_FILL[str(value).lower()])
                cell.font = Font(bold=True, color=WHITE)
    if rows:
        try:
            ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
            tab = Table(displayName=_safe_table_name(ws.title + str(start_row)), ref=ref)
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            ws.add_table(tab)
        except Exception:
            pass


def ROW_FILL():
    return CARD_ALT


def _paint(ws, rows=40, cols=10):
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=BG)


def _header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, color=RED_DARK)
        c.fill = PatternFill("solid", fgColor=CARD_ALT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format(ws, freeze="A5"):
    border = Border(left=Side(style="thin", color=BORDER), right=Side(style="thin", color=BORDER), top=Side(style="thin", color=BORDER), bottom=Side(style="thin", color=BORDER))
    for row in ws.iter_rows():
        max_height = 18
        for cell in row:
            cell.border = border
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                max_height = max(max_height, min(95, 16 + len(str(cell.value)) // 55 * 11))
        ws.row_dimensions[row[0].row].height = max_height
    for column_cells in ws.columns:
        length = 0
        col = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), 62))
        ws.column_dimensions[col].width = max(13, min(48, length + 2))
    if freeze:
        ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False


def _excel_value(value):
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def _safe_table_name(title):
    safe = "".join(ch for ch in str(title).title() if ch.isalnum())[:24]
    return safe or "OpenEasmTable"


def _loc(loc):
    if not isinstance(loc, dict):
        return ""
    return loc.get("display") or loc.get("path") or loc.get("record") or loc.get("hostname") or loc.get("control") or ""


def _conclusion(audit):
    return (
        f"Le domaine {audit.get('domain')} obtient un score de {audit.get('score', {}).get('score')} / 1000. "
        f"Profil : {audit.get('domain_profile', {}).get('label', 'N/A')}. "
        f"Surface observée : {audit.get('ip_inventory', {}).get('public_ip_count', 0)} IP publiques, "
        f"{audit.get('subdomains', {}).get('count', 0)} sous-domaines, "
        f"{audit.get('service_scan', {}).get('count_open_ports', 0)} ports ouverts et "
        f"{audit.get('service_scan', {}).get('count_cves', 0)} CVE service/version."
    )


def _severity_order(sev):
    return {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}.get(str(sev).lower(), 5)


def _priority_label(sev):
    return {"critical": "P1 immédiat", "high": "P2 prioritaire", "medium": "P3 planifié", "low": "P4 amélioration", "info": "Information"}.get(str(sev).lower(), "Information")


def _sla_for(sev):
    return {"critical": "< 5 jours", "high": "< 15 jours", "medium": "< 30 jours", "low": "< 90 jours", "info": "Suivi"}.get(str(sev).lower(), "Suivi")
