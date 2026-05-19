from __future__ import annotations

import io
import json
import os
import socket
import time
from pathlib import Path
from datetime import datetime, timezone

import dns.resolver
import httpx
from sqlalchemy import text
from sqlalchemy.orm import Session

REPORT_DIR = Path("/app/reports")

def _item(name: str, status: str, message: str, details: dict | None = None) -> dict:
    return {
        "name": name,
        "status": status,
        "message": message,
        "details": details or {},
    }

def _overall(items: list[dict]) -> str:
    if any(i["status"] == "error" for i in items):
        return "error"
    if any(i["status"] == "warning" for i in items):
        return "warning"
    return "ok"

def check_database(db: Session) -> dict:
    try:
        db.execute(text("SELECT 1"))
        return _item("PostgreSQL", "ok", "Connexion base de données opérationnelle.")
    except Exception as exc:
        return _item("PostgreSQL", "error", "Connexion base de données impossible.", {"error": str(exc)})

def check_reports_dir() -> dict:
    try:
        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        test = REPORT_DIR / ".openeasm_write_test"
        test.write_text("ok", encoding="utf-8")
        test.unlink(missing_ok=True)
        return _item("Dossier reports", "ok", "Le dossier des rapports existe et est accessible en écriture.", {"path": str(REPORT_DIR)})
    except Exception as exc:
        return _item("Dossier reports", "error", "Le dossier des rapports n'est pas accessible en écriture.", {"error": str(exc), "path": str(REPORT_DIR)})

def check_dns() -> dict:
    try:
        resolver = dns.resolver.Resolver()
        resolver.lifetime = 3
        resolver.timeout = 3
        answers = resolver.resolve("google.com", "A")
        values = [str(a) for a in answers][:5]
        return _item("Résolution DNS", "ok", "Résolution DNS opérationnelle.", {"sample": values})
    except Exception as exc:
        return _item("Résolution DNS", "warning", "Résolution DNS dégradée ou indisponible.", {"error": str(exc)})

def check_outbound_https() -> dict:
    try:
        with httpx.Client(timeout=5.0, follow_redirects=True) as client:
            response = client.get("https://crt.sh/?q=%.example.com&output=json")
        if response.status_code < 500:
            return _item("Source passive crt.sh", "ok", f"crt.sh répond avec HTTP {response.status_code}.")
        return _item("Source passive crt.sh", "warning", f"crt.sh répond avec HTTP {response.status_code}. L'audit continuera avec les sources alternatives.")
    except Exception as exc:
        return _item("Source passive crt.sh", "warning", "crt.sh inaccessible depuis le conteneur. Les sources alternatives et fallbacks restent utilisés.", {"error": str(exc)})

def test_export_dependencies(cleanup: bool = True) -> dict:
    created_files = []
    checks = []

    try:
        REPORT_DIR.mkdir(parents=True, exist_ok=True)

        # JSON
        json_path = REPORT_DIR / f"openeasm_export_test_{int(time.time())}.json"
        json_path.write_text(json.dumps({"status": "ok", "ts": datetime.now(timezone.utc).isoformat()}), encoding="utf-8")
        created_files.append(json_path)
        checks.append(_item("Export JSON", "ok", "Écriture JSON opérationnelle."))

        # Excel
        try:
            from openpyxl import Workbook
            xlsx_path = REPORT_DIR / f"openeasm_export_test_{int(time.time())}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Export Test"
            ws["A1"] = "OpenEASM"
            ws["B1"] = "OK"
            wb.save(xlsx_path)
            created_files.append(xlsx_path)
            checks.append(_item("Export Excel", "ok", "Génération Excel opérationnelle."))
        except Exception as exc:
            checks.append(_item("Export Excel", "error", "Génération Excel impossible.", {"error": str(exc)}))

        # PDF
        try:
            from reportlab.pdfgen import canvas
            pdf_path = REPORT_DIR / f"openeasm_export_test_{int(time.time())}.pdf"
            c = canvas.Canvas(str(pdf_path))
            c.drawString(72, 720, "OpenEASM export test OK")
            c.save()
            created_files.append(pdf_path)
            checks.append(_item("Export PDF", "ok", "Génération PDF opérationnelle."))
        except Exception as exc:
            checks.append(_item("Export PDF", "error", "Génération PDF impossible.", {"error": str(exc)}))

    finally:
        if cleanup:
            for path in created_files:
                try:
                    path.unlink(missing_ok=True)
                except Exception:
                    pass

    return {
        "overall": _overall(checks),
        "items": checks,
        "created_files": [p.name for p in created_files],
        "cleanup": cleanup,
    }

def build_system_diagnostics(db: Session) -> dict:
    items = [
        _item("API FastAPI", "ok", "API OpenEASM opérationnelle."),
        check_database(db),
        check_reports_dir(),
        check_dns(),
        check_outbound_https(),
    ]

    export_test = test_export_dependencies(cleanup=True)
    items.extend(export_test["items"])

    return {
        "version": "alpha",
        "app": "OpenEASM",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "overall": _overall(items),
        "items": items,
        "environment": {
            "python": os.sys.version.split()[0],
            "hostname": socket.gethostname(),
            "reports_dir": str(REPORT_DIR),
        },
    }
